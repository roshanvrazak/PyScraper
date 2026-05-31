# -*- coding: utf-8 -*-
"""
Scrapy Item Pipelines
----------------------
Cleans, normalizes, deduplicates and exports data to a professionally styled
Excel spreadsheet using openpyxl.
"""

import os
import logging
from datetime import datetime
from scrapy.exceptions import DropItem

logger = logging.getLogger(__name__)

# Soft import openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.error("openpyxl is required for Excel export. Install using: pip install openpyxl")
    openpyxl = None


class ExcelExportPipeline:
    """Cleans scraped jobs, deduplicates, and saves them into an elegant Excel sheet."""
    
    def __init__(self):
        self.seen_urls = set()
        self.items = []
        self.output_filename = "curated_sponsored_jobs.xlsx"
        
    def open_spider(self, spider):
        logger.info("ExcelExportPipeline opened. Initializing collector...")
        
    def process_item(self, item, spider):
        # 1. Clean application URL and check for duplicates
        app_url = item.get("application_url", "").strip()
        if not app_url:
            raise DropItem("Missing application URL. Dropping.")
            
        if app_url.lower() in self.seen_urls:
            raise DropItem(f"Duplicate job detected by URL: {app_url}")
            
        self.seen_urls.add(app_url.lower())
        
        # 2. Clean & normalize data fields
        item["company_name"] = item.get("company_name", "").strip()
        item["job_title"] = item.get("job_title", "").strip()
        item["location"] = item.get("location", "").strip()
        
        salary = item.get("salary_range")
        if not salary or str(salary).strip() in ["", "None", "null"]:
            item["salary_range"] = "Not Specified"
        else:
            item["salary_range"] = str(salary).strip()
            
        # Standardize boolean exclusion flag
        item["sponsorship_keyword_found"] = bool(item.get("sponsorship_keyword_found", False))
        
        # Add scraping date
        if not item.get("date_scraped"):
            item["date_scraped"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
        self.items.append(item)
        return item
        
    def close_spider(self, spider):
        if not openpyxl:
            logger.error("ExcelExportPipeline: openpyxl not found. Skipping file output.")
            return
            
        if not self.items:
            logger.warning("ExcelExportPipeline: No items scraped. Excel file will not be created.")
            return
            
        logger.info(f"ExcelExportPipeline: Processing {len(self.items)} jobs for Excel export.")
        
        # Sort items: Group by company name first, then by job title
        self.items.sort(key=lambda x: (x["company_name"].lower(), x["job_title"].lower()))
        
        # Create a new workbook and select active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sponsored Tech Jobs"
        
        # Enable visible gridlines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # Design Palette & Fonts
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)
        warning_font = Font(name=font_family, size=10, color="9C0006", bold=True)
        ok_font = Font(name=font_family, size=10, color="006100")
        
        # Curated Fills (Steel Navy for headers, zebra coloring for rows)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # Borders
        thin_border_side = Side(style='thin', color='D9D9D9')
        cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Alignments
        left_align = Alignment(horizontal='left', vertical='center')
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Column headers definitions
        headers = [
            "Company Name", 
            "Job Title", 
            "Location", 
            "Salary Range", 
            "Visa Exclusions Found?", 
            "Application URL", 
            "Date Scraped"
        ]
        
        # Write headers
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = cell_border
            
        # Write records
        for row_idx, item in enumerate(self.items, 2):
            # Check if visa/sponsorship exclusions were flagged
            excl_flag = item["sponsorship_keyword_found"]
            excl_text = "Yes" if excl_flag else "No"
            
            # Retrieve cells
            c_company = ws.cell(row=row_idx, column=1, value=item["company_name"])
            c_title = ws.cell(row=row_idx, column=2, value=item["job_title"])
            c_loc = ws.cell(row=row_idx, column=3, value=item["location"])
            c_salary = ws.cell(row=row_idx, column=4, value=item["salary_range"])
            c_excl = ws.cell(row=row_idx, column=5, value=excl_text)
            c_url = ws.cell(row=row_idx, column=6, value=item["application_url"])
            c_date = ws.cell(row=row_idx, column=7, value=item["date_scraped"])
            
            # Apply base fonts and alignments
            for cell in [c_company, c_title, c_loc, c_salary, c_url]:
                cell.font = data_font
                cell.alignment = left_align
                cell.border = cell_border
                
            c_date.font = data_font
            c_date.alignment = center_align
            c_date.border = cell_border
            
            # Apply exclusion cell specific styling
            c_excl.alignment = center_align
            c_excl.border = cell_border
            if excl_flag:
                c_excl.font = warning_font
                c_excl.fill = warning_fill
            else:
                c_excl.font = ok_font
                c_excl.fill = ok_fill
                
            # Subtle Zebra striping for even rows
            if row_idx % 2 == 0:
                for cell in [c_company, c_title, c_loc, c_salary, c_url, c_date]:
                    cell.fill = zebra_fill
                    
        # Apply Auto-Filter
        max_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{max_col_letter}{len(self.items) + 1}"
        
        # Auto-adjust column widths dynamically based on length of contents
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row == 1:
                    val_str += "  "  # Add padding for excel filter arrows
                max_len = max(max_len, len(val_str))
            
            col_letter = get_column_letter(col[0].column)
            # Limit extremely long URLs from stretching column indefinitely
            adjusted_width = min(max(max_len + 3, 12), 45)
            ws.column_dimensions[col_letter].width = adjusted_width
            
        # Freeze headers so they remain visible when scrolling
        ws.freeze_panes = "A2"
        
        # Save output
        try:
            wb.save(self.output_filename)
            logger.info(f"ExcelExportPipeline: Successfully exported Excel sheet: '{os.path.abspath(self.output_filename)}'")
        except Exception as e:
            logger.error(f"ExcelExportPipeline: Failed to save Excel file '{self.output_filename}': {e}")
