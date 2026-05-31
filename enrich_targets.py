#!/usr/bin/env python3
"""
UK Gov Sponsorship Licence CSV → Career Pages Excel Enrichment
--------------------------------------------------------------
Reads the UK Government register of licensed sponsors (CSV), filters for
tech/CS-relevant companies, resolves their career page URLs asynchronously,
and writes a styled Excel workbook ready for downstream job-scraping agents.

Usage:
    python3 enrich_targets.py --file sponsor_licence_register.csv --limit 100
    python3 enrich_targets.py --file sponsor_licence_register.csv --output my_targets.xlsx
"""

import os
import sys
import asyncio
import argparse
import logging
import random
import re
from datetime import datetime
from typing import Optional
from urllib.parse import urlparse

try:
    import pandas as pd
except ImportError:
    print("[ERROR] pandas is required. Run: pip install pandas", file=sys.stderr)
    sys.exit(1)

try:
    import httpx
except ImportError:
    print("[ERROR] httpx is required. Run: pip install httpx", file=sys.stderr)
    sys.exit(1)

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("[ERROR] beautifulsoup4 is required. Run: pip install beautifulsoup4", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("enrich_targets.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("enrich_targets")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Keywords used to identify tech/CS-relevant companies by name.
# The UK Gov CSV has no sector column so we match on Organisation Name.
TECH_KEYWORDS = [
    # Core tech identifiers
    "software", "tech", "technology", "technologies",
    "computer", "computing", "cyber",
    # Data / AI / ML
    "artificial intelligence", "machine learning", "deep learning",
    "data science", "data engineering", "analytics",
    # Cloud / infrastructure
    "cloud", "devops", "saas", "paas", "iaas",
    # Specific tech domains
    "semiconductor", "robotics", "automation",
    "fintech", "edtech", "healthtech", "proptech", "legaltech",
    "internet", "platform",
    # IT services (specific enough)
    "it services", "it consulting", "it solutions",
    "digital transformation",
]

# URL path fragments that indicate a genuine career/jobs page
CAREER_URL_SIGNALS = [
    "career", "careers", "jobs", "job", "work-with-us", "work-for-us",
    "join-us", "join-our-team", "opportunities", "vacancies", "hiring",
    "recruitment", "talent",
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

# ---------------------------------------------------------------------------
# Argument parsing
# ---------------------------------------------------------------------------

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Enrich UK Gov sponsor CSV with career page URLs → Excel output."
    )
    parser.add_argument(
        "--file", "-f",
        default="dummy_sponsors.csv",
        help="Path to the UK Gov sponsor licence CSV (or xlsx). Default: dummy_sponsors.csv",
    )
    parser.add_argument(
        "--output", "-o",
        default="career_targets.xlsx",
        help="Output Excel file path. Default: career_targets.xlsx",
    )
    parser.add_argument(
        "--limit", "-l",
        type=int,
        default=None,
        help="Process only the first N filtered companies (useful for testing).",
    )
    parser.add_argument(
        "--concurrency", "-c",
        type=int,
        default=2,
        help="Max concurrent HTTP requests (keep ≤3 to avoid rate-limiting). Default: 2",
    )
    parser.add_argument(
        "--delay", "-d",
        type=float,
        default=2.0,
        help="Base delay in seconds between requests. Default: 2.0",
    )
    parser.add_argument(
        "--no-filter",
        action="store_true",
        help="Skip tech keyword filtering and process ALL companies in the file.",
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# CSV loading & filtering
# ---------------------------------------------------------------------------

def load_and_filter(file_path: str, no_filter: bool) -> pd.DataFrame:
    """Load the UK Gov sponsor CSV and optionally filter for tech companies."""
    if not os.path.exists(file_path):
        logger.error(f"File not found: {file_path}")
        sys.exit(1)

    ext = os.path.splitext(file_path)[1].lower()
    logger.info(f"Reading: {file_path}")

    try:
        if ext == ".csv":
            df = pd.read_csv(file_path, dtype=str)
        elif ext in (".xlsx", ".xls"):
            df = pd.read_excel(file_path, dtype=str)
        else:
            logger.error(f"Unsupported format '{ext}'. Use .csv, .xls, or .xlsx")
            sys.exit(1)
    except Exception as e:
        logger.error(f"Failed to read file: {e}")
        sys.exit(1)

    df = df.fillna("").apply(lambda col: col.str.strip() if col.dtype == "object" else col)
    logger.info(f"Loaded {len(df)} rows.")

    # Detect the organisation name column
    name_col = _detect_name_column(df)
    logger.info(f"Organisation name column detected: '{name_col}'")

    # Drop blank / duplicate names
    df = df[df[name_col].str.len() > 0].drop_duplicates(subset=[name_col])

    if no_filter:
        logger.info("--no-filter set: keeping all companies.")
        return df, name_col

    # Filter by tech keywords (case-insensitive, match anywhere in name)
    pattern = "|".join(re.escape(kw) for kw in TECH_KEYWORDS)
    mask = df[name_col].str.contains(pattern, case=False, na=False)
    filtered = df[mask].copy()
    logger.info(
        f"Tech filter: {len(filtered)} / {len(df)} companies match "
        f"({len(df) - len(filtered)} excluded)."
    )
    return filtered, name_col


def _detect_name_column(df: pd.DataFrame) -> str:
    candidates = [
        "organisation name", "organization name", "company name",
        "company", "organisation", "organization", "name",
    ]
    for col in df.columns:
        if col.strip().lower() in candidates:
            return col
    return df.columns[0]


# ---------------------------------------------------------------------------
# Resume: load previously enriched rows from output Excel
# ---------------------------------------------------------------------------

def load_existing_results(output_path: str) -> dict:
    """Returns a dict of {company_name: row_dict} from a previous run's Excel."""
    if not os.path.exists(output_path):
        return {}
    try:
        existing_df = pd.read_excel(output_path, dtype=str).fillna("")
        if "Organisation Name" not in existing_df.columns:
            return {}
        results = {}
        for _, row in existing_df.iterrows():
            name = row.get("Organisation Name", "").strip()
            if name:
                results[name] = row.to_dict()
        logger.info(f"Resuming: {len(results)} companies already in '{output_path}'.")
        return results
    except Exception as e:
        logger.warning(f"Could not load existing output for resume: {e}. Starting fresh.")
        return {}


# ---------------------------------------------------------------------------
# Career URL discovery
# ---------------------------------------------------------------------------

def _confidence(career_url: Optional[str]) -> str:
    """Score the career URL quality: High / Medium / Low / Not Found."""
    if not career_url:
        return "Not Found"
    path = urlparse(career_url).path.lower() + career_url.lower()
    if any(sig in path for sig in CAREER_URL_SIGNALS):
        return "High"
    # URL found but no obvious career signal — could be homepage
    return "Medium"


async def _ddg_search(client: httpx.AsyncClient, query: str, delay: float) -> Optional[str]:
    """Query DuckDuckGo HTML interface and return the first external result URL."""
    await asyncio.sleep(delay + random.uniform(0.3, 1.5))

    search_url = "https://html.duckduckgo.com/html/"
    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-GB,en;q=0.8",
        "Referer": "https://duckduckgo.com/",
    }

    backoff = 2.0
    for attempt in range(1, 4):
        try:
            resp = await client.get(
                search_url,
                params={"q": query},
                headers=headers,
                timeout=15.0,
            )

            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")

                # Primary: result__url links
                for tag in soup.find_all("a", class_="result__url"):
                    href = _extract_ddg_href(tag.get("href", ""))
                    if href:
                        return href

                # Fallback: result__a (title links) which carry uddg param
                for tag in soup.find_all("a", class_="result__a"):
                    href = _extract_ddg_href(tag.get("href", ""))
                    if href:
                        return href

                return None

            if resp.status_code in (429, 503):
                logger.warning(f"DDG rate-limited (attempt {attempt}). Waiting {backoff}s…")
                await asyncio.sleep(backoff)
                backoff *= 2
            else:
                logger.error(f"DDG returned HTTP {resp.status_code} for query: {query}")
                return None

        except (httpx.RequestError, asyncio.TimeoutError) as exc:
            logger.warning(f"DDG request error (attempt {attempt}): {exc}")
            if attempt == 3:
                return None
            await asyncio.sleep(backoff)
            backoff *= 2

    return None


def _extract_ddg_href(href: str) -> Optional[str]:
    """Decode a DuckDuckGo result href into a plain external URL."""
    if not href:
        return None
    if "/l/?uddg=" in href or "uddg=" in href:
        from urllib.parse import parse_qs, unquote
        qs = href.split("?", 1)[-1]
        params = parse_qs(qs)
        uddg = params.get("uddg", [None])[0]
        if uddg:
            href = unquote(uddg)
    domain = urlparse(href).netloc.lower()
    if domain and "duckduckgo.com" not in domain:
        return href
    return None


async def _serpapi_search(client: httpx.AsyncClient, query: str, api_key: str) -> Optional[str]:
    """Use SerpAPI (Google) for higher-quality results when an API key is set."""
    try:
        resp = await client.get(
            "https://serpapi.com/search",
            params={"engine": "google", "q": query, "api_key": api_key, "num": 3},
            timeout=10.0,
        )
        if resp.status_code == 200:
            organic = resp.json().get("organic_results", [])
            if organic:
                return organic[0].get("link")
        else:
            logger.error(f"SerpAPI error {resp.status_code}: {resp.text[:200]}")
    except Exception as exc:
        logger.error(f"SerpAPI request failed: {exc}")
    return None


async def resolve_career_url(
    client: httpx.AsyncClient,
    company_name: str,
    semaphore: asyncio.Semaphore,
    delay: float,
    serpapi_key: Optional[str],
) -> Optional[str]:
    """Find the best career page URL for a company. Returns URL or None."""
    async with semaphore:
        logger.info(f"Searching: '{company_name}'")
        query = f'"{company_name}" UK careers jobs site'

        url = None
        if serpapi_key:
            url = await _serpapi_search(client, query, serpapi_key)

        if not url:
            url = await _ddg_search(client, query, delay)

        if url:
            logger.info(f"  Found [{_confidence(url)}]: {url}")
        else:
            logger.warning(f"  Not found: '{company_name}'")

        return url


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

# Column order in output Excel — mirrors UK Gov CSV columns + enrichment cols
OUTPUT_COLUMNS = [
    "Organisation Name",
    "Town/City",
    "County",
    "Type & Rating",
    "Route",
    "Career Page URL",
    "Confidence",
    "Status",
    "Date Added",
]

# Styling constants
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
FOUND_FILL = PatternFill("solid", fgColor="C6EFCE")
FOUND_FONT = Font(color="276221", bold=True)
NOT_FOUND_FILL = PatternFill("solid", fgColor="FFC7CE")
NOT_FOUND_FONT = Font(color="9C0006", bold=True)
MEDIUM_FILL = PatternFill("solid", fgColor="FFEB9C")
MEDIUM_FONT = Font(color="9C5700", bold=True)
ZEBRA_FILL = PatternFill("solid", fgColor="F2F6FA")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7E0"),
    right=Side(style="thin", color="D0D7E0"),
    top=Side(style="thin", color="D0D7E0"),
    bottom=Side(style="thin", color="D0D7E0"),
)


def write_excel(rows: list[dict], output_path: str):
    """Write enriched results to a styled Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Career Targets"

    # Header row
    ws.append(OUTPUT_COLUMNS)
    for col_idx, _ in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        zebra = row_idx % 2 == 0
        status = row.get("Status", "Not Found")
        confidence = row.get("Confidence", "Not Found")

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Colour-code the Confidence / Status columns
            if col_name in ("Confidence", "Status"):
                if confidence == "High" or status == "Found":
                    cell.fill = FOUND_FILL
                    cell.font = FOUND_FONT
                elif confidence == "Medium":
                    cell.fill = MEDIUM_FILL
                    cell.font = MEDIUM_FONT
                else:
                    cell.fill = NOT_FOUND_FILL
                    cell.font = NOT_FOUND_FONT
            elif zebra:
                cell.fill = ZEBRA_FILL

        # Make Career Page URL a clickable hyperlink
        url_cell = ws.cell(row=row_idx, column=OUTPUT_COLUMNS.index("Career Page URL") + 1)
        if url_cell.value and url_cell.value.startswith("http"):
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="1155CC", underline="single")

    # Auto-fit column widths
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = str(row[0].value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Freeze header row and enable auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    total = len(rows)
    found = sum(1 for r in rows if r.get("Status") == "Found")
    high = sum(1 for r in rows if r.get("Confidence") == "High")
    medium = sum(1 for r in rows if r.get("Confidence") == "Medium")
    not_found = total - found

    summary_data = [
        ("Total Companies Processed", total),
        ("Career URLs Found", found),
        ("High Confidence URLs", high),
        ("Medium Confidence URLs", medium),
        ("Not Found", not_found),
        ("Success Rate", f"{(found/total*100):.1f}%" if total else "0%"),
        ("Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        lbl_cell = summary_ws.cell(row=r_idx, column=1, value=label)
        val_cell = summary_ws.cell(row=r_idx, column=2, value=value)
        lbl_cell.font = Font(bold=True)
        lbl_cell.fill = PatternFill("solid", fgColor="EBF3FB")
        val_cell.fill = PatternFill("solid", fgColor="F8FBFE")

    summary_ws.column_dimensions["A"].width = 32
    summary_ws.column_dimensions["B"].width = 20

    wb.save(output_path)
    logger.info(f"Excel saved → {output_path}  ({total} rows, {found} URLs found)")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

async def run(args: argparse.Namespace):
    # Load source CSV
    filtered_df, name_col = load_and_filter(args.file, args.no_filter)

    # Resume: load already-enriched companies
    existing = load_existing_results(args.output)

    # Determine which companies still need lookup
    all_companies = filtered_df[name_col].tolist()
    pending = [c for c in all_companies if c not in existing]
    logger.info(f"{len(existing)} already resolved, {len(pending)} pending.")

    if args.limit:
        pending = pending[: args.limit]
        logger.info(f"--limit {args.limit}: processing {len(pending)} companies this run.")

    serpapi_key = os.environ.get("SERPAPI_API_KEY")
    if serpapi_key:
        logger.info("SERPAPI_API_KEY found — using Google Search engine.")
    else:
        logger.info("No SERPAPI_API_KEY — using DuckDuckGo HTML fallback.")

    # Build a lookup map for original row metadata
    meta_map = {
        row[name_col]: row.to_dict()
        for _, row in filtered_df.iterrows()
    }

    # Async enrichment
    semaphore = asyncio.Semaphore(args.concurrency)

    async with httpx.AsyncClient(follow_redirects=True) as client:
        tasks = {
            company: resolve_career_url(client, company, semaphore, args.delay, serpapi_key)
            for company in pending
        }
        task_results = await asyncio.gather(*tasks.values(), return_exceptions=True)
        resolved = dict(zip(tasks.keys(), task_results))

    # Merge new results into existing
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for company, result in resolved.items():
        url = result if isinstance(result, str) else None
        confidence = _confidence(url)
        meta = meta_map.get(company, {})

        existing[company] = {
            "Organisation Name": company,
            "Town/City": meta.get("Town/City", ""),
            "County": meta.get("County", ""),
            "Type & Rating": meta.get("Type & Rating", ""),
            "Route": meta.get("Route", ""),
            "Career Page URL": url or "",
            "Confidence": confidence,
            "Status": "Found" if url else "Not Found",
            "Date Added": now,
        }

    # Also ensure previously enriched companies have all output columns
    for company, row in existing.items():
        if "Status" not in row:
            row["Status"] = "Found" if row.get("Career Page URL") else "Not Found"
        if "Confidence" not in row:
            row["Confidence"] = _confidence(row.get("Career Page URL"))

    # Write Excel
    rows = list(existing.values())
    write_excel(rows, args.output)


def main():
    args = parse_arguments()
    try:
        asyncio.run(run(args))
    except KeyboardInterrupt:
        logger.info("\nInterrupted. Progress is saved in the output Excel.")


if __name__ == "__main__":
    main()
